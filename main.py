#!/usr/bin/env python3
"""
Sanitization App v1.0
"""
import os, re, uuid, shutil, tempfile, logging, subprocess, argparse, sys, io
import platform, stat, urllib.request, tarfile, zipfile

from datetime import datetime, timedelta
from threading import Semaphore, Thread, Lock
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, send_file, render_template_string
from logging.handlers import RotatingFileHandler
from cryptography import x509
from cryptography.x509.oid import NameOID
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import rsa

# --- Config ---
def _on_pythonanywhere():
    """True when running on PythonAnywhere (free or paid)."""
    return bool(os.environ.get('PYTHONANYWHERE_DOMAIN') or os.environ.get('PYTHONANYWHERE_SITE'))


# Local/dev default 2 GB; PythonAnywhere free cannot accept that over HTTP.
MAX_UPLOAD_MB = 100 if _on_pythonanywhere() else 2048
MAX_THREADS   = 4 if _on_pythonanywhere() else 15
MAX_FILES     = 100 if _on_pythonanywhere() else 10000
PURGE_DAYS    = 1           # Retention days for uploads/logs/outputs/UUIDs
LARGE_FILE_MB = 50          # Use line-by-line sanitization above this size
ARCHIVE_EXTS = {
    '.zip','.rar','.7z','.tar','.tgz','.tar.gz',
    '.bz2','.tar.bz2','.gz','.xz','.lz','.zst'
}
COMPOUND_ARCHIVE_SUFFIXES = ('.tar.gz', '.tar.bz2', '.tar.xz', '.tbz2', '.tgz')
TEXT_EXTS = {'.txt', '.log', '.csv'}
SPREADSHEET_EXTS = {'.xlsx', '.xls'}

RULES_FILE = 'bad_words.txt'

# --- Log Rotation and Directories Setup ---
BASE   = Path(__file__).parent.resolve()
UPLOAD = BASE / 'uploads'
OUTPUT = BASE / 'outputs'
LOGS   = BASE / 'logs'
TEMP   = BASE / 'temp'
TOOLS  = BASE / 'tools' / '7zip'
for d in (UPLOAD, OUTPUT, LOGS, TEMP):
    d.mkdir(exist_ok=True)


def max_upload_label():
    """Human-readable upload size for UI (e.g. '100 MB' or '2 GB')."""
    if MAX_UPLOAD_MB >= 1024 and MAX_UPLOAD_MB % 1024 == 0:
        return f"{MAX_UPLOAD_MB // 1024} GB"
    return f"{MAX_UPLOAD_MB} MB"

SEVEN_ZIP_EXE = None
SEVEN_ZIP_VERSION = '2409'

log_handler = RotatingFileHandler(
    LOGS / 'processing.log',
    maxBytes=5*1024*1024,
    backupCount=3
)
logging.basicConfig(
    handlers=[log_handler],
    level=logging.INFO,
    format="%(asctime)s %(levelname)s:%(message)s"
)

rules_lock = Lock()


def x_replacer(match):
    """Replace matched text with X repeated for each character in the match."""
    return 'X' * len(match.group(0))


def ensure_selfsigned_certs():
    """Ensure development self-signed certificates are present."""
    CRT_DIR = BASE / 'certs'
    CRT_DIR.mkdir(exist_ok=True)
    crt = CRT_DIR / 'cert.pem'
    key = CRT_DIR / 'key.pem'

    if crt.exists() and key.exists():
        return

    logging.info("Generating dev self-signed cert/key using cryptography.")

    try:
        key_obj = rsa.generate_private_key(
            public_exponent=65537,
            key_size=2048,
        )

        subject = issuer = x509.Name([
            x509.NameAttribute(NameOID.COMMON_NAME, u"SanitizationAppDevCert"),
        ])

        cert = x509.CertificateBuilder()\
            .subject_name(subject)\
            .issuer_name(issuer)\
            .public_key(key_obj.public_key())\
            .serial_number(x509.random_serial_number())\
            .not_valid_before(datetime.utcnow())\
            .not_valid_after(datetime.utcnow() + timedelta(days=3650))\
            .sign(key_obj, hashes.SHA256())

        with open(key, "wb") as f:
            f.write(key_obj.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.TraditionalOpenSSL,
                encryption_algorithm=serialization.NoEncryption()
            ))

        with open(crt, "wb") as f:
            f.write(cert.public_bytes(serialization.Encoding.PEM))

    except Exception as e:
        logging.error("Could not generate cert.pem/key.pem: %s", e)
        print("Could not generate cert.pem/key.pem:", e, file=sys.stderr)
        sys.exit(1)


def _download_file(url, dest, verbose=False):
    """Download a file from url to dest."""
    logging.info("Downloading %s", url)
    if verbose:
        print(f"  Downloading: {url}", flush=True)
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    urllib.request.urlretrieve(url, dest)
    if verbose:
        print(f"  Download saved to: {dest}", flush=True)


def _mark_executable(path):
    """Ensure a binary is executable on Unix systems."""
    path.chmod(path.stat().st_mode | stat.S_IXUSR | stat.S_IXGRP | stat.S_IXOTH)


def _find_7z_in_tree(root):
    """Search for a 7-Zip binary under root."""
    if not root.exists():
        return None
    for name in ('7z.exe', '7za.exe', '7zz', '7za', '7z'):
        for candidate in root.rglob(name):
            if candidate.is_file():
                if name.endswith('.exe') or platform.system() != 'Windows':
                    if platform.system() != 'Windows':
                        _mark_executable(candidate)
                    return candidate
    return None


def find_7z():
    """Locate an existing 7-Zip executable."""
    bundled = _find_7z_in_tree(TOOLS)
    if bundled:
        return bundled
    for name in ('7z', '7za', '7zz'):
        found = shutil.which(name)
        if found:
            return Path(found)
    if platform.system() == 'Windows':
        for base in (Path(r'C:\Program Files\7-Zip'), Path(r'C:\Program Files (x86)\7-Zip')):
            exe = base / '7z.exe'
            if exe.exists():
                return exe
    return None


def _is_ubuntu():
    """True on Ubuntu and Ubuntu derivatives (e.g. Pop!_OS)."""
    try:
        data = Path('/etc/os-release').read_text(encoding='utf-8')
    except OSError:
        return False
    fields = {}
    for line in data.splitlines():
        if '=' in line:
            k, v = line.split('=', 1)
            fields[k.strip()] = v.strip().strip('"')
    os_id = fields.get('ID', '').lower()
    id_like = fields.get('ID_LIKE', '').lower()
    return os_id == 'ubuntu' or 'ubuntu' in id_like.split()





def _verify_7zip(exe):
    """Run 7-Zip to confirm it works. Returns (ok, version_line)."""
    try:
        result = subprocess.run(
            [str(exe)],
            capture_output=True,
            text=True,
            timeout=15,
        )
        output = (result.stdout or '') + (result.stderr or '')
        version = None
        for line in output.splitlines()[:8]:
            stripped = line.strip()
            if stripped and ('7-Zip' in stripped or '7-zip' in stripped.lower()):
                version = stripped
                break
        if not version and output.strip():
            version = output.strip().splitlines()[0].strip()
        return True, version
    except (OSError, subprocess.SubprocessError, subprocess.TimeoutExpired) as e:
        logging.error("7-Zip verification failed for %s: %s", exe, e)
        return False, None


def _install_7zip_winget(verbose=False):
    """Install 7-Zip on Windows via winget."""
    if not shutil.which('winget'):
        if verbose:
            print("  winget not available on this system.", flush=True)
        return None
    if verbose:
        print("  Installing 7-Zip via winget (please wait)...", flush=True)
    logging.info("Installing 7-Zip via winget...")
    result = subprocess.run(
        [
            'winget', 'install', '--id', '7zip.7zip', '-e',
            '--silent', '--accept-package-agreements', '--accept-source-agreements',
        ],
        capture_output=True,
        text=True,
    )
    if verbose and result.stdout.strip():
        print(f"  {result.stdout.strip()}", flush=True)
    if result.returncode != 0 and verbose:
        print(f"  winget exit code: {result.returncode}", flush=True)
    return find_7z()




def _install_7zip_apt(verbose=False):
    """Install p7zip on Debian/Ubuntu via apt (may require sudo)."""
    if not shutil.which('apt-get'):
        if verbose:
            print("  apt-get not available on this system.", flush=True)
        return None
    if verbose:
        print("  Installing p7zip-full via apt (sudo password may be required)...", flush=True)
    logging.info("Installing p7zip via apt-get (sudo may be required)...")
    result = subprocess.run(
        ['sudo', 'apt-get', 'install', '-y', 'p7zip-full'],
        text=True,
    )
    if result.returncode != 0 and verbose:
        print(f"  apt install failed (exit code {result.returncode}).", flush=True)
    return find_7z()


def _install_7zip_dnf(verbose=False):
    """Install p7zip on RHEL/Fedora via dnf or yum."""
    for manager, packages in (
        ('dnf', ['p7zip', 'p7zip-plugins']),
        ('yum', ['p7zip']),
    ):
        if not shutil.which(manager):
            continue
        if verbose:
            print(f"  Installing p7zip via {manager} (sudo password may be required)...", flush=True)
        result = subprocess.run(
            ['sudo', manager, 'install', '-y', *packages],
            text=True,
        )
        if result.returncode == 0:
            return find_7z()
        if verbose:
            print(f"  {manager} install failed (exit code {result.returncode}).", flush=True)
    return None


def _install_7zip_portable_windows(verbose=False):
    """Download portable 7-Zip extra package for Windows (no admin required)."""
    if verbose:
        print("  Installing portable 7-Zip (no admin required)...", flush=True)
    TOOLS.mkdir(parents=True, exist_ok=True)
    seven_zr = TOOLS / '7zr.exe'
    extra = TOOLS / f'7z{SEVEN_ZIP_VERSION}-extra.7z'
    if not seven_zr.exists():
        _download_file('https://www.7-zip.org/a/7zr.exe', seven_zr, verbose=verbose)
    if not extra.exists():
        _download_file(
            f'https://www.7-zip.org/a/7z{SEVEN_ZIP_VERSION}-extra.7z',
            extra,
            verbose=verbose,
        )
    seven_za = TOOLS / '7za.exe'
    if not seven_za.exists():
        if verbose:
            print("  Extracting portable 7-Zip files...", flush=True)
        subprocess.run(
            [str(seven_zr), 'x', str(extra), f'-o{TOOLS}', '-y'],
            check=True,
            capture_output=not verbose,
        )
    return _find_7z_in_tree(TOOLS)


def _install_7zip_portable_linux(verbose=False):
    """Download portable 7-Zip binary for Linux (no admin required)."""
    if verbose:
        print("  Installing portable 7-Zip for Linux (no admin required)...", flush=True)
    TOOLS.mkdir(parents=True, exist_ok=True)
    machine = platform.machine().lower()
    arch = 'linux-arm64' if machine in ('aarch64', 'arm64') else 'linux-x64'
    tarball = TOOLS / f'7z{SEVEN_ZIP_VERSION}-{arch}.tar.xz'
    if not tarball.exists():
        _download_file(
            f'https://www.7-zip.org/a/7z{SEVEN_ZIP_VERSION}-{arch}.tar.xz',
            tarball,
            verbose=verbose,
        )
    if verbose:
        print("  Extracting portable 7-Zip...", flush=True)
    with tarfile.open(tarball, 'r:xz') as tar:
        tar.extractall(TOOLS)
    exe = _find_7z_in_tree(TOOLS)
    if exe:
        _mark_executable(exe)
    return exe


def _print_7zip_help():
    """Print manual install instructions for the current platform."""
    system = platform.system()
    print("", flush=True)
    print("  Automatic installation was attempted but did not complete.", flush=True)
    print("  Text file uploads will still work. Archive uploads need 7-Zip.", flush=True)
    print("", flush=True)
    print("  Manual steps if needed:", flush=True)
    if system == 'Windows':
        print("  - Restart the app (portable 7-Zip is downloaded to tools/7zip/)", flush=True)
        print("  - Or install from https://www.7-zip.org/", flush=True)
    elif system == 'Darwin':
        print("  - Restart the app to retry automatic + p7zip install", flush=True)
        
    elif _is_ubuntu():
        print("  - Restart the app to retry automatic apt install", flush=True)
        print("  - Or run: sudo apt-get update && sudo apt-get install -y p7zip-full", flush=True)
    else:
        print("  - Restart the app to retry portable 7-Zip download", flush=True)
        print("  - Or run: sudo apt-get install -y p7zip-full", flush=True)


def ensure_7zip(verbose=False):
    """Ensure 7-Zip is available; auto-install if missing."""
    global SEVEN_ZIP_EXE

    if SEVEN_ZIP_EXE and Path(SEVEN_ZIP_EXE).exists():
        return Path(SEVEN_ZIP_EXE)

    if verbose:
        print("", flush=True)
        print("=" * 54, flush=True)
        print("  7-Zip Setup", flush=True)
        print("=" * 54, flush=True)
        print("  Checking for 7-Zip...", flush=True)

    found = find_7z()
    if found:
        ok, version = _verify_7zip(found)
        if ok:
            SEVEN_ZIP_EXE = found
            logging.info("Using 7-Zip: %s", found)
            if verbose:
                print("  Status: OK — 7-Zip is ready", flush=True)
                print(f"  Path:   {found}", flush=True)
                if version:
                    print(f"  Info:   {version}", flush=True)
                print("=" * 54, flush=True)
                print("", flush=True)
            return found
        if verbose:
            print(f"  Found {found} but it did not run correctly.", flush=True)
            print("  Will attempt automatic installation...", flush=True)
    elif verbose:
        print("  Status: Not found on this system.", flush=True)
        print("  Attempting automatic installation...", flush=True)

    logging.info("7-Zip not found. Attempting automatic installation...")
    system = platform.system()
    exe = None
    method = None
    on_pa = _on_pythonanywhere()

    try:
        if system == 'Windows':
            if shutil.which('winget'):
                method = 'winget'
                exe = False
            if not exe:
                method = 'portable download (no admin)'
                exe = _install_7zip_portable_windows(verbose=verbose)
        elif system == 'Darwin':
            method = ''
            exe = False
        elif on_pa or (system == 'Linux' and not _is_ubuntu()):
            # PythonAnywhere free has no sudo; use portable 7-Zip first.
            method = 'portable Linux download (no admin)'
            exe = _install_7zip_portable_linux(verbose=verbose)
            if not exe and shutil.which('apt-get') and not on_pa:
                method = 'apt (p7zip-full)'
                exe = _install_7zip_apt(verbose=verbose)
            if not exe and not on_pa:
                method = 'dnf/yum (p7zip)'
                exe = _install_7zip_dnf(verbose=verbose)
        elif _is_ubuntu():
            method = 'apt (p7zip-full)'
            exe = _install_7zip_apt(verbose=verbose)
            if not exe:
                method = 'portable Linux download (no admin)'
                exe = _install_7zip_portable_linux(verbose=verbose)
        else:
            method = 'portable Linux download (no admin)'
            exe = _install_7zip_portable_linux(verbose=verbose)
            if not exe and shutil.which('apt-get'):
                method = 'apt (p7zip-full)'
                exe = _install_7zip_apt(verbose=verbose)
            if not exe:
                method = 'dnf/yum (p7zip)'
                exe = _install_7zip_dnf(verbose=verbose)

    except Exception as e:
        logging.error("Automatic 7-Zip installation failed: %s", e)
        if verbose:
            print(f"  ERROR during install ({method}): {e}", flush=True)

    if exe and exe.exists():
        ok, version = _verify_7zip(exe)
        if ok:
            SEVEN_ZIP_EXE = exe
            logging.info("7-Zip ready: %s", exe)
            if verbose:
                print("", flush=True)
                print("  Status: OK — 7-Zip installed successfully", flush=True)
                print(f"  Method: {method}", flush=True)
                print(f"  Path:   {exe}", flush=True)
                if version:
                    print(f"  Info:   {version}", flush=True)
                print("=" * 54, flush=True)
                print("", flush=True)
            return exe
        if verbose:
            print(f"  Installed at {exe} but verification failed.", flush=True)

    logging.warning("7-Zip could not be installed automatically.")
    SEVEN_ZIP_EXE = None
    if verbose:
        print("", flush=True)
        print("  Status: FAILED — Could not install 7-Zip automatically", flush=True)
        if method:
            print(f"  Last method tried: {method}", flush=True)
        _print_7zip_help()
        print("=" * 54, flush=True)
        print("", flush=True)
    else:
        print(
            "Warning: 7-Zip is not available. Archive uploads will not work.",
            file=sys.stderr,
        )
    return None


def get_7z_exe():
    """Return the 7-Zip executable path, installing if needed."""
    if SEVEN_ZIP_EXE and Path(SEVEN_ZIP_EXE).exists():
        return Path(SEVEN_ZIP_EXE)
    return ensure_7zip()


def get_archive_ext(name):
    """Return archive extension, including compound suffixes like .tar.gz."""
    lower = name.lower()
    for ext in COMPOUND_ARCHIVE_SUFFIXES:
        if lower.endswith(ext):
            return ext
    return Path(name).suffix.lower()


def is_archive_name(name):
    """Check if a filename is a supported archive."""
    return get_archive_ext(name) in ARCHIVE_EXTS


def rules_path():
    return BASE / RULES_FILE


def read_rules_file():
    """Return raw lines from bad_words.txt (including comments and blanks)."""
    path = rules_path()
    if not path.exists():
        return []
    return path.read_text('utf-8').splitlines()


def parse_active_patterns(lines):
    """Extract active regex patterns, skipping comments and empty lines."""
    patterns = []
    for line in lines:
        clean = line.split('#', 1)[0].strip()
        if clean:
            patterns.append(clean)
    return patterns


def validate_patterns(patterns):
    """Validate regex patterns. Returns (ok, error_message)."""
    for i, pat in enumerate(patterns, start=1):
        try:
            re.compile(pat)
        except re.error as e:
            return False, f"Line {i}: invalid regex — {e}"
    return True, None


def validate_patterns_in_lines(lines):
    """Validate active regex patterns using real file line numbers."""
    for i, line in enumerate(lines, start=1):
        clean = line.split('#', 1)[0].strip()
        if not clean:
            continue
        try:
            re.compile(clean)
        except re.error as e:
            return False, f"Line {i}: invalid regex — {e}"
    return True, None


def write_rules_file(lines):
    """Atomically write rules file and reload global patterns."""
    global BAD_PATTERNS
    path = rules_path()
    tmp = path.with_suffix('.txt.tmp')
    content = '\n'.join(lines)
    if lines:
        content += '\n'
    with rules_lock:
        tmp.write_text(content, encoding='utf-8')
        tmp.replace(path)
        BAD_PATTERNS = parse_active_patterns(lines)
    return BAD_PATTERNS


def load_bad_patterns():
    """Load patterns from bad_words.txt."""
    return parse_active_patterns(read_rules_file())


def reload_bad_patterns():
    """Reload patterns from disk."""
    global BAD_PATTERNS
    with rules_lock:
        BAD_PATTERNS = load_bad_patterns()
    return BAD_PATTERNS


def purge_old():
    """Purge old uploads, logs, outputs, and expired sessions."""
    cutoff = datetime.now() - timedelta(days=PURGE_DAYS)
    for folder in (LOGS, OUTPUT, UPLOAD):
        for f in folder.iterdir():
            if f.is_file() and datetime.fromtimestamp(f.stat().st_mtime) < cutoff:
                try:
                    f.unlink()
                except OSError:
                    pass
    purge_ids = [
        k for k, v in SESSIONS.items()
        if 'time' in v and v['time'] < cutoff
    ]
    for k in purge_ids:
        SESSIONS.pop(k, None)


def is_text(fp: Path) -> bool:
    """Determine if a file is a text file (by attempting UTF-8 decode)."""
    try:
        with open(fp, 'r', encoding='utf-8') as f:
            f.read(4096)
        return True
    except (UnicodeDecodeError, OSError):
        return False


def sanitize_content(text, rules=None):
    """Sanitize content by replacing non-ASCII and bad patterns."""
    repl_total = 0
    text, n = re.subn(r"[^\x20-\x7E\n\r]", x_replacer, text)
    repl_total += n
    pats = rules if rules is not None else BAD_PATTERNS
    for pat in pats:
        text, n = re.subn(pat, x_replacer, text, flags=re.IGNORECASE)
        repl_total += n
    return text, repl_total


def sanitize_text_file(fpath, rules=None):
    """Sanitize a text file in-place, using streaming for large files."""
    size_mb = fpath.stat().st_size / (1024 * 1024)
    if size_mb <= LARGE_FILE_MB:
        with open(fpath, 'r+', encoding='utf-8', errors='ignore') as f:
            text = f.read()
            text, n = sanitize_content(text, rules)
            f.seek(0)
            f.write(text)
            f.truncate()
        return n

    total = 0
    tmp_out = fpath.with_suffix(fpath.suffix + '.sanit.tmp')
    with open(fpath, 'r', encoding='utf-8', errors='ignore') as src, \
         open(tmp_out, 'w', encoding='utf-8') as dst:
        for line in src:
            sanitized, n = sanitize_content(line, rules)
            dst.write(sanitized)
            total += n
    tmp_out.replace(fpath)
    return total


def sanitize_xlsx_file(fpath, rules=None):
    """Sanitize string cell values in an .xlsx workbook in-place."""
    import openpyxl
    wb = openpyxl.load_workbook(fpath)
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    new_val, n = sanitize_content(cell.value, rules)
                    cell.value = new_val
                    total += n
    wb.save(fpath)
    return total


def sanitize_xls_file(fpath, rules=None):
    """Sanitize string cell values in a legacy .xls workbook in-place."""
    import xlrd
    import xlwt
    rb = xlrd.open_workbook(fpath)
    wb = xlwt.Workbook()
    total = 0
    for sheet_idx in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_idx)
        ws = wb.add_sheet(rs.name)
        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                val = rs.cell_value(row_idx, col_idx)
                if isinstance(val, str):
                    val, n = sanitize_content(val, rules)
                    total += n
                ws.write(row_idx, col_idx, val)
    wb.save(str(fpath))
    return total


def sanitize_file_if_supported(fpath: Path, rules=None) -> int:
    """Sanitize file contents when the extension or encoding is supported."""
    ext = fpath.suffix.lower()
    if ext == '.xlsx':
        return sanitize_xlsx_file(fpath, rules)
    if ext == '.xls':
        return sanitize_xls_file(fpath, rules)
    if ext in TEXT_EXTS or is_text(fpath):
        return sanitize_text_file(fpath, rules)
    return 0


def sanitize_filename_windows(name):
    """Replace Windows-forbidden filename chars and all non-ASCII characters."""
    bad_pattern = r'[<>:"/\\|?*\x00-\x1F]|[^\x20-\x7E]'
    sanitized = re.sub(bad_pattern, x_replacer, name)
    sanitized = sanitized.rstrip(' .')
    if not sanitized or sanitized in ('.', '..'):
        sanitized = 'file.txt'
    num_replacements = len(re.findall(bad_pattern, name))
    return sanitized, num_replacements


def is_junk_mac_file(path):
    """Check if path is a macOS junk file/folder."""
    s = str(path)
    return (
        '/__MACOSX' in s or '\\__MACOSX' in s or
        s.endswith('.DS_Store') or s.endswith('/.DS_Store') or s.endswith('\\.DS_Store')
    )


def safe_rename_in_dir(path, used_names):
    """
    Rename file or directory to sanitized version, avoiding collisions in its directory.
    Returns: (Path to new file, sanitized name, number of filename replacements)
    """
    dir_path = path.parent
    safe, fname_repl = sanitize_filename_windows(path.name)
    candidate = dir_path / safe
    orig_safe = safe
    counter = 1
    while (candidate.exists() and candidate != path) or safe in used_names:
        stem, ext = os.path.splitext(orig_safe)
        safe = f"{stem}_{counter}{ext}"
        candidate = dir_path / safe
        counter += 1
    if candidate != path:
        try:
            path.rename(candidate)
            logging.info("Renamed: %s -> %s", path.name, candidate.name)
        except FileNotFoundError:
            logging.warning("Could not rename (not found): %s", path)
            return path, path.name, 0
        except Exception as e:
            logging.error("Rename failed: %s -> %s (%s)", path, candidate, e)
            return path, path.name, 0
    used_names.add(safe)
    return candidate, safe, fname_repl


def extract_archive(upload_path, tempdir):
    """Extract an archive to a temporary directory using 7z, with stdlib fallback."""
    upload_path = Path(upload_path)
    tempdir = Path(tempdir)
    tempdir.mkdir(parents=True, exist_ok=True)

    exe = get_7z_exe()
    if exe:
        subprocess.run(
            [str(exe), 'x', str(upload_path), f'-o{tempdir}', '-y'],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return

    # Fallback for hosts without 7-Zip (e.g. PythonAnywhere free outbound limits)
    lower = upload_path.name.lower()
    try:
        if lower.endswith('.zip'):
            with zipfile.ZipFile(upload_path, 'r') as zf:
                zf.extractall(tempdir)
            return
        if lower.endswith(('.tar', '.tar.gz', '.tgz', '.tar.bz2', '.tbz2', '.tar.xz')):
            with tarfile.open(upload_path, 'r:*') as tf:
                tf.extractall(tempdir)
            return
        shutil.unpack_archive(str(upload_path), str(tempdir))
        return
    except Exception as e:
        logging.error(
            "Archive extraction failed (no 7-Zip and stdlib fallback failed) for %s: %s",
            upload_path, e,
        )


def sanitize_tree(root):
    """Recursively sanitize all files and directory names in a directory tree."""
    filename_replacements = 0

    for dirpath, dirnames, filenames in os.walk(root, topdown=False):
        path = Path(dirpath)
        used_names = set()
        for fname in filenames:
            f = path / fname
            if is_junk_mac_file(f):
                try:
                    f.unlink()
                except OSError:
                    pass
                continue
            if is_archive_name(fname):
                extract_archive(f, f.parent)
                try:
                    f.unlink()
                except OSError:
                    pass
                filename_replacements += sanitize_tree(f.parent)
                continue
            _, _, frepl = safe_rename_in_dir(f, used_names)
            filename_replacements += frepl
        for dname in dirnames:
            d = path / dname
            if is_junk_mac_file(d):
                try:
                    shutil.rmtree(d)
                except OSError:
                    pass
                continue
            _, _, drepl = safe_rename_in_dir(d, used_names)
            filename_replacements += drepl
    return filename_replacements


def repack(tmpdir: Path, out_base: Path, ext: str) -> Path:
    """Repack the directory as a new archive based on extension."""
    e = ext.lower()
    if e == '.bz2':
        files = [p for p in tmpdir.iterdir() if p.is_file()]
        if len(files) == 1:
            out = out_base.with_suffix('.bz2')
            subprocess.run(f"bzip2 -c '{files[0]}' > '{out}'", shell=True, check=True)
            return out
        fmt = 'bztar'
    elif e in ('.tar.bz2', '.tbz2'):
        fmt = 'bztar'
    elif e in ('.tar.gz', '.tgz', '.gz'):
        fmt = 'gztar'
    elif e == '.tar':
        fmt = 'tar'
    else:
        fmt = 'zip'
    archive = shutil.make_archive(str(out_base), fmt, root_dir=tmpdir)
    return Path(archive)


# --- Flask & Processing App ---

app = Flask(__name__)
app.secret_key = uuid.uuid4().hex[:8]
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

SESSIONS = {}
sem = Semaphore(MAX_THREADS)

BAD_PATTERNS = load_bad_patterns()


def process_file(job_id, upload_paths, names, original_names, ip, rules=None):
    """Process each uploaded file/archive."""
    sem.acquire()
    try:
        SESSIONS[job_id]['status'] = 'Processing'
        active_rules = rules if rules is not None else reload_bad_patterns()
        tmp = Path(tempfile.mkdtemp(dir=TEMP))

        sanitized_names = []
        per_upload_stats = []
        output_paths = []
        total_global_replacements = 0
        total_global_duration = 0.0

        for idx, up in enumerate(upload_paths):
            orig = original_names[idx]
            safe, fname_repl = sanitize_filename_windows(orig)
            candidate = tmp / safe
            counter = 1
            orig_safe = safe
            while candidate.exists():
                stem, ext = os.path.splitext(orig_safe)
                safe = f"{stem}_{counter}{ext}"
                candidate = tmp / safe
                counter += 1

            arch_ext = get_archive_ext(safe)
            replacements = fname_repl
            start = datetime.now()
            if is_archive_name(safe):
                extract_archive(up, tmp)
                replacements += sanitize_tree(tmp)
                for dirpath, dirnames, filenames in os.walk(tmp):
                    for fname in filenames:
                        fpath = Path(dirpath) / fname
                        replacements += sanitize_file_if_supported(fpath, active_rules)
                out_base = OUTPUT / f"sanit_{job_id}_{Path(safe).stem}"
                out = repack(tmp, out_base, arch_ext)
            else:
                shutil.copy(up, candidate)
                replacements += sanitize_file_if_supported(candidate, active_rules)
                out = OUTPUT / f"{job_id}_{safe}"
                shutil.copy2(candidate, out)
            duration = (datetime.now() - start).total_seconds()
            sanitized_names.append(safe)
            per_upload_stats.append({
                'file': safe,
                'replacements': replacements,
                'duration': round(duration, 5)
            })
            output_paths.append(str(out))
            total_global_replacements += replacements
            total_global_duration += duration

        allow_download = any(stat['replacements'] > 0 for stat in per_upload_stats)

        session_update = {
            'status': 'Done',
            'file_stats': per_upload_stats,
            'total_replacements': total_global_replacements,
            'duration': total_global_duration,
            'output_path': output_paths[0] if output_paths else "",
            'time': datetime.now()
        }
        if allow_download:
            session_update['download'] = f"/download/{job_id}?token={SESSIONS[job_id]['token']}"
        else:
            session_update['download'] = None

        SESSIONS[job_id].update(session_update)

        log_file = LOGS / f"{job_id}.log"
        with open(log_file, 'w') as lf:
            lf.write(",".join([
                job_id,
                ";".join(sanitized_names),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                ip,
                f"{total_global_duration:.2f}",
                str(total_global_replacements)
            ]) + "\n")

        shutil.rmtree(tmp, ignore_errors=True)
        for up in upload_paths:
            up.unlink(missing_ok=True)
    finally:
        sem.release()


@app.errorhandler(413)
def handle_413(e):
    return jsonify(error=f'File too large (max {MAX_UPLOAD_MB}MB)'), 413


@app.route('/')
def index():
    html = (BASE / 'templates' / 'index.html').read_text('utf-8')
    return render_template_string(
        html,
        max_mb=MAX_UPLOAD_MB,
        max_files=MAX_FILES,
        max_label=max_upload_label(),
        year=datetime.now().year,
    )


@app.route('/logo.png', strict_slashes=False)
def logo():
    """Serve the logo image or a minimal placeholder."""
    logo_path = BASE / 'templates' / 'logo.png'
    if logo_path.exists():
        return send_file(logo_path, mimetype='image/png')
    svg = (
        b'<svg xmlns="http://www.w3.org/2000/svg" width="60" height="60" viewBox="0 0 60 60">'
        b'<rect width="60" height="60" rx="8" fill="#F97316"/>'
        b'<text x="30" y="38" font-family="Arial,sans-serif" font-size="18" fill="#fff" '
        b'text-anchor="middle" font-weight="bold">SOC</text></svg>'
    )
    return send_file(
        io.BytesIO(svg),
        mimetype='image/svg+xml'
    )


@app.route('/upload', methods=['POST'])
def upload():
    """Handle upload request, save files and start processing thread."""
    ensure_7zip(verbose=False)
    purge_old()
    files = request.files.getlist('file')
    if not files or len(files) > MAX_FILES:
        return jsonify(error=f"Select 1-{MAX_FILES} files"), 400

    job_id = uuid.uuid4().hex[:8]
    SESSIONS[job_id] = {
        'status': 'Queued',
        'token': uuid.uuid4().hex,
        'time': datetime.now()
    }

    upaths, names, original_names = [], [], []
    for f in files:
        fn = f.filename
        up = UPLOAD / f"{job_id}_{fn}"
        f.save(up)
        upaths.append(up)
        names.append(fn)
        original_names.append(f.filename)

    Thread(
        target=process_file,
        args=(job_id, upaths, names, original_names, request.remote_addr, None),
        daemon=True
    ).start()

    return jsonify(job_id=job_id)


@app.route('/status/<job_id>')
def status(job_id):
    """Return the current status for the given job."""
    info = SESSIONS.get(job_id)
    if info:
        resp = {
            'job_id': job_id,
            'status': info['status'],
            'progress': info.get('progress')
        }
        if info['status'] == 'Done':
            resp.update({
                'download': info['download'],
                'file_stats': info['file_stats'],
                'duration': info['duration'],
                'total_replacements': info['total_replacements']
            })
        return jsonify(resp)

    lf = LOGS / f"{job_id}.log"
    if not lf.exists():
        return jsonify(error='Invalid Job ID'), 404
    parts = lf.read_text('utf-8').strip().split(',')
    _, names_semi, ts, ip, dur, repl = parts
    return jsonify({
        'job_id': job_id,
        'status': 'Done',
        'download': f"/download/{job_id}",
        'duration': float(dur),
        'total_replacements': int(repl),
        'file_stats': []
    })


@app.route('/download/<job_id>')
def download(job_id):
    token = request.args.get("token", "")
    info = SESSIONS.get(job_id)
    if not info or info.get("token") != token or info.get("status") != "Done":
        return ('Forbidden', 403)

    out = Path(info['output_path'])
    if not out.exists():
        return ('Not Found', 404)

    download_name = out.name
    if info['file_stats'] and len(info['file_stats']) == 1:
        download_name = info['file_stats'][0]['file']

    return send_from_directory(
        out.parent, out.name, as_attachment=True, download_name=download_name
    )


@app.route('/bad_words')
def bad_words():
    """Serve all lines from the rules file."""
    lines = read_rules_file()
    return jsonify(lines=lines)


@app.route('/rules/save', methods=['POST'])
def rules_save():
    """Replace entire rules file."""
    data = request.get_json(silent=True) or {}
    lines = data.get('lines')
    if lines is None:
        return jsonify(error='Missing lines'), 400
    if not isinstance(lines, list):
        return jsonify(error='lines must be a list'), 400

    ok, err = validate_patterns_in_lines(lines)
    if not ok:
        return jsonify(error=err), 400

    write_rules_file(lines)
    return jsonify(status='Rules saved', count=len(parse_active_patterns(lines)))


@app.route('/rules/add', methods=['POST'])
def rules_add():
    """Append one rule to the rules file."""
    data = request.get_json(silent=True) or {}
    rule = (data.get('rule') or '').strip()
    if not rule:
        return jsonify(error='Rule cannot be empty'), 400

    ok, err = validate_patterns([rule])
    if not ok:
        return jsonify(error=err), 400

    lines = read_rules_file()
    lines.append(rule)
    write_rules_file(lines)
    return jsonify(status='Rule added', count=len(parse_active_patterns(lines)))


@app.route('/rules/remove', methods=['POST'])
def rules_remove():
    """Remove a rule by line index."""
    data = request.get_json(silent=True) or {}
    index = data.get('index')
    if index is None or not isinstance(index, int):
        return jsonify(error='Missing or invalid index'), 400

    lines = read_rules_file()
    if index < 0 or index >= len(lines):
        return jsonify(error='Index out of range'), 400

    removed = lines.pop(index)
    write_rules_file(lines)
    return jsonify(status='Rule removed', removed=removed)


@app.route('/rules/edit', methods=['POST'])
def rules_edit():
    """Update a single line by index."""
    data = request.get_json(silent=True) or {}
    index = data.get('index')
    line = data.get('line')
    if index is None or not isinstance(index, int):
        return jsonify(error='Missing or invalid index'), 400
    if line is None:
        return jsonify(error='Missing line content'), 400

    lines = read_rules_file()
    if index < 0 or index >= len(lines):
        return jsonify(error='Index out of range'), 400

    clean = line.split('#', 1)[0].strip()
    if clean:
        ok, err = validate_patterns_in_lines([line])
        if not ok:
            return jsonify(error=err.replace('Line 1:', f'Line {index + 1}:')), 400

    lines[index] = line
    write_rules_file(lines)
    return jsonify(status='Line updated', count=len(parse_active_patterns(lines)))


@app.route('/rules/clear', methods=['POST'])
def rules_clear():
    """Clear all rules."""
    write_rules_file([])
    return jsonify(status='All rules cleared')


@app.route('/rules/upload', methods=['POST'])
def rules_upload():
    """Upload a .txt file to replace all rules."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify(error='No file uploaded'), 400

    try:
        content = f.read().decode('utf-8')
    except UnicodeDecodeError:
        return jsonify(error='File must be UTF-8 text'), 400

    lines = content.splitlines()
    ok, err = validate_patterns_in_lines(lines)
    if not ok:
        return jsonify(error=err), 400

    write_rules_file(lines)
    return jsonify(status='Rules file uploaded and replaced', count=len(parse_active_patterns(lines)))


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--serve', action='store_true', help='Run HTTPS server')
    parser.add_argument('file', nargs='?', help='CLI input file')
    parser.add_argument('--rules', nargs='*', help='Inline regex rules')
    parser.add_argument('--rules-file', help='Load regex rules from file')
    args = parser.parse_args()

    ensure_selfsigned_certs()

    if args.serve:
        debug = True
        # Flask debug reloader runs this script twice; only show startup in the server process
        if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not debug:
            ensure_7zip(verbose=True)
            print("Starting Sanitization App...", flush=True)
            print("Open https://localhost:8443 in your browser (accept the security warning).", flush=True)
            print("Press Ctrl+C to stop.", flush=True)
            print("", flush=True)
        app.run(
            host='0.0.0.0',
            port=8443,
            ssl_context=(str(BASE / 'certs' / 'cert.pem'), str(BASE / 'certs' / 'key.pem')),
            threaded=True,
            debug=debug
        )
    elif args.file:
        ensure_7zip(verbose=False)
        purge_old()
        f = Path(args.file)
        if not f.is_file():
            print("Error: file not found", file=sys.stderr)
            sys.exit(1)

        job_id = uuid.uuid4().hex[:8]
        SESSIONS[job_id] = {
            'status': 'Queued',
            'token': uuid.uuid4().hex,
            'time': datetime.now()
        }
        up = UPLOAD / f"{job_id}_{f.name}"
        shutil.copy(f, up)

        safe, nrep = sanitize_filename_windows(f.name)
        rules = None
        if args.rules_file:
            rules = parse_active_patterns(
                Path(args.rules_file).read_text('utf-8').splitlines()
            )
        elif args.rules:
            rules = args.rules

        process_file(job_id, [up], [safe], [f.name], 'CLI', rules)
        info = SESSIONS[job_id]
        print("Job ID:", job_id)
        print(f"Duration: {info.get('duration', 0):.2f}s")
        print(f"Total Replacements: {info.get('total_replacements', 0)}")
        if info.get('total_replacements', 0) > 0:
            print(f"Output file: {info.get('output_path', 'N/A')}")
        else:
            print("No replacements found. No output file generated.")
    else:
        parser.print_help()
